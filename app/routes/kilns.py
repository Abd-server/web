"""
═══════════════════════════════════════════════════════════════
  مسارات الأفران (kilns.py) — محدّث للحقول الحقيقية
═══════════════════════════════════════════════════════════════
"""
from __future__ import annotations

from typing import List

from fastapi import APIRouter, Depends, HTTPException, status, Body
from sqlalchemy.orm import Session

from app.core.deps import get_current_user
from app.core.device_auth import get_kiln_by_device_key
from app.core.notifications import process_reading_notifications, send_notification, log_event
from app.db.database import get_db
from app.models.user import User
from app.models.kiln import Kiln, Reading, Event, generate_device_key
from app.models.kiln_schemas import (
    KilnCreate, KilnUpdate, KilnResponse, KilnWithKeyResponse,
    ReadingIngest, ReadingResponse,
    NotifySettings, NotifySettingsResponse, StopStatusResponse,
    EventResponse,
)

from datetime import timedelta, timezone as _tz
try:
    from zoneinfo import ZoneInfo
except ImportError:
    ZoneInfo = None

def _local_time(dt, tz_name="Asia/Muscat"):
    """يحوّل وقت UTC إلى المنطقة الزمنية المحددة ويُرجّعه نصاً، أو '' لو فارغ."""
    if not dt:
        return ""
    # الأوقات تُحفظ بـ UTC (بدون tzinfo)، نُلحق UTC ثم نحوّل
    if ZoneInfo is not None:
        try:
            aware = dt.replace(tzinfo=_tz.utc)
            return aware.astimezone(ZoneInfo(tz_name)).strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    # احتياطي: عُمان +4
    return (dt + timedelta(hours=4)).strftime("%Y-%m-%d %H:%M:%S")

router = APIRouter(tags=["الأفران"])


def _get_owned_kiln(kiln_id: str, user: User, db: Session) -> Kiln:
    kiln = (
        db.query(Kiln)
        .filter(Kiln.id == kiln_id, Kiln.owner_id == user.id)
        .first()
    )
    if kiln is None:
        raise HTTPException(status_code=404, detail="الفرن غير موجود")
    return kiln


# ═════════════ مسارات المستخدم ═════════════

@router.post("/kilns", response_model=KilnWithKeyResponse, status_code=201)
def create_kiln(body: KilnCreate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = Kiln(owner_id=current_user.id, name=body.name, location=body.location)
    db.add(kiln); db.commit(); db.refresh(kiln)
    return kiln


@router.get("/kilns", response_model=List[KilnResponse])
def list_my_kilns(current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return (
        db.query(Kiln).filter(Kiln.owner_id == current_user.id)
        .order_by(Kiln.created_at.desc()).all()
    )


@router.get("/kilns/{kiln_id}", response_model=KilnResponse)
def get_kiln(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    return _get_owned_kiln(kiln_id, current_user, db)


@router.patch("/kilns/{kiln_id}", response_model=KilnResponse)
def update_kiln(kiln_id: str, body: KilnUpdate, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    for field, value in body.dict(exclude_unset=True).items():
        setattr(kiln, field, value)
    db.commit(); db.refresh(kiln)
    return kiln


@router.delete("/kilns/{kiln_id}", status_code=204)
def delete_kiln(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    db.delete(kiln); db.commit()
    return None


@router.get("/kilns/{kiln_id}/readings", response_model=List[ReadingResponse])
def get_kiln_readings(kiln_id: str, limit: int = 200, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_owned_kiln(kiln_id, current_user, db)
    limit = max(1, min(limit, 2000))
    rows = (
        db.query(Reading).filter(Reading.kiln_id == kiln_id)
        .order_by(Reading.recorded_at.desc()).limit(limit).all()
    )
    return list(reversed(rows))  # أقدم→أحدث (مناسب للرسم البياني)


@router.get("/kilns/{kiln_id}/latest", response_model=ReadingResponse)
def get_latest_reading(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    _get_owned_kiln(kiln_id, current_user, db)
    reading = (
        db.query(Reading).filter(Reading.kiln_id == kiln_id)
        .order_by(Reading.recorded_at.desc()).first()
    )
    if reading is None:
        raise HTTPException(status_code=404, detail="لا توجد قراءات بعد")
    return reading


@router.get("/kilns/{kiln_id}/status")
def get_connection_status(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """حالة اتصال الفرن: متصل لو وصلت قراءة خلال آخر دقيقتين."""
    from datetime import datetime, timedelta
    _get_owned_kiln(kiln_id, current_user, db)
    last = (
        db.query(Reading).filter(Reading.kiln_id == kiln_id)
        .order_by(Reading.recorded_at.desc()).first()
    )
    if last is None or last.recorded_at is None:
        return {"connected": False, "last_seen": None}
    elapsed = (datetime.utcnow() - last.recorded_at).total_seconds()
    return {"connected": elapsed <= 120, "last_seen": last.recorded_at.isoformat()}


@router.post("/cron/check-offline")
def cron_check_offline(db: Session = Depends(get_db)):
    """
    فحص دوري: يكتشف الأفران التي انقطعت (لم ترسل خلال 3 دقائق) ويُشعر أصحابها مرة واحدة.
    يُستدعى من مؤقت خارجي (cron) كل بضع دقائق. لا يتطلب مصادقة مستخدم.
    """
    from datetime import datetime, timedelta
    from app.core.notifications import send_notification
    cutoff = datetime.utcnow() - timedelta(minutes=3)
    notified = 0
    # نفحص فقط الأفران التي كانت متصلة (was_online==1)
    kilns = db.query(Kiln).filter(Kiln.was_online == 1).all()
    for kiln in kilns:
        last = (
            db.query(Reading).filter(Reading.kiln_id == kiln.id)
            .order_by(Reading.recorded_at.desc()).first()
        )
        if last is None or last.recorded_at is None:
            continue
        if last.recorded_at < cutoff:
            # انقطع
            try:
                send_notification(kiln, "⚠️ انقطع اتصال الفرن بالإنترنت. لم تصل قراءات منذ أكثر من 3 دقائق.", title="انقطاع الاتصال", db=db)
                kiln.was_online = 0
                db.commit()
                notified += 1
            except Exception as e:
                print(f"تنبيه: تعذّر إشعار الانقطاع: {e}")
    return {"checked": len(kilns), "offline_notified": notified}


@router.get("/kilns/{kiln_id}/device-key", response_model=KilnWithKeyResponse)
def get_device_key(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """يرجّع مفتاح الجهاز للمالك (يُعرض من السيرفر، فيظهر على أي متصفح/جهاز)."""
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    return kiln


@router.post("/kilns/{kiln_id}/telegram/subscribe-code")
def create_kiln_tg_code(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """ينشئ رمز اشتراك تيليجرام لهذا الفرن — أي شخص يمسحه يصبح مشتركاً في إشعاراته."""
    from app.core.config import settings
    from app.core.telegram_bot import generate_kiln_link_code
    if not settings.telegram_configured():
        raise HTTPException(status_code=503, detail="بوت تيليجرام غير مُهيّأ")
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    code = generate_kiln_link_code(kiln, db)
    return {"code": code, "expires_minutes": 10, "bot_username": "KilnMonitor_bot",
            "deep_link": f"https://t.me/KilnMonitor_bot?start={code}"}


@router.get("/kilns/{kiln_id}/telegram/subscribers")
def list_kiln_tg_subscribers(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """قائمة مشتركي تيليجرام لهذا الفرن."""
    from app.models.kiln import TelegramSubscriber
    _get_owned_kiln(kiln_id, current_user, db)
    subs = db.query(TelegramSubscriber).filter(TelegramSubscriber.kiln_id == kiln_id).all()
    return [{"id": s.id, "name": s.name or "مشترك", "created_at": s.created_at.isoformat() if s.created_at else None} for s in subs]


@router.delete("/kilns/{kiln_id}/telegram/subscribers/{sub_id}")
def remove_kiln_tg_subscriber(kiln_id: str, sub_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """يحذف مشترك تيليجرام من الفرن."""
    from app.models.kiln import TelegramSubscriber
    _get_owned_kiln(kiln_id, current_user, db)
    sub = db.query(TelegramSubscriber).filter(TelegramSubscriber.id == sub_id, TelegramSubscriber.kiln_id == kiln_id).first()
    if sub:
        db.delete(sub); db.commit()
    return {"message": "تم حذف المشترك"}


@router.post("/kilns/{kiln_id}/rotate-key", response_model=KilnWithKeyResponse)
def rotate_device_key(kiln_id: str, body: dict = Body(default={}), current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    # حماية: يجب إرسال تأكيد صريح (confirm=="تجديد") لمنع الضغط بالخطأ
    if str(body.get("confirm", "")).strip() != "تجديد":
        raise HTTPException(status_code=400, detail="يلزم تأكيد التجديد")
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    kiln.device_key = generate_device_key()
    db.commit(); db.refresh(kiln)
    return kiln


# ═════════════ مسار الجهاز (الأردوينو) ═════════════

@router.post("/device/readings", response_model=ReadingResponse, status_code=201)
def ingest_reading(body: ReadingIngest, kiln: Kiln = Depends(get_kiln_by_device_key), db: Session = Depends(get_db)):
    """
    يستقبل قراءة من الأردوينو. الفرن يُحدَّد من X-Device-Key.
    - يُخزّن قراءة واحدة كل دقيقة (يتجاهل الأكثر تكراراً لتقليل حجم البيانات).
    - يحذف القراءات الأقدم من 6 أشهر تلقائياً (تنظيف دوري خفيف).
    - الأحداث (تحوّل المرحلة/الإيقاف) تُسجّل دائماً عبر معالجة الإشعارات، حتى لو لم تُخزّن القراءة.
    """
    from datetime import datetime, timezone as _tz, timedelta

    reading = Reading(kiln_id=kiln.id, **body.dict(exclude_unset=True))

    # هل مرّت دقيقة على آخر قراءة مخزّنة؟
    last = (
        db.query(Reading).filter(Reading.kiln_id == kiln.id)
        .order_by(Reading.recorded_at.desc()).first()
    )
    store_it = True
    if last and last.recorded_at:
        elapsed = datetime.utcnow() - last.recorded_at
        if elapsed < timedelta(seconds=58):   # أقل من دقيقة → لا نخزّن
            store_it = False

    saved_reading = last  # افتراضياً نُرجّع الأخيرة لو لم نخزّن
    if store_it:
        db.add(reading); db.commit(); db.refresh(reading)
        saved_reading = reading

        # تنظيف دوري: حذف الأقدم من 6 أشهر (يُنفّذ أحياناً فقط لتقليل الحمل)
        try:
            import random
            if random.random() < 0.02:   # ~2% من الكتابات تُشغّل التنظيف
                cutoff = datetime.utcnow() - timedelta(days=183)
                db.query(Reading).filter(
                    Reading.kiln_id == kiln.id, Reading.recorded_at < cutoff
                ).delete(synchronize_session=False)
                db.commit()
        except Exception as e:
            print(f"تنبيه: تعذّر تنظيف القراءات القديمة: {e}")

    # معالجة الإشعارات دائماً (حتى لو لم نخزّن القراءة) — لرصد تحوّل المرحلة والإيقاف
    try:
        process_reading_notifications(kiln, reading, db)
    except Exception as e:
        print(f"تنبيه: تعذّرت معالجة الإشعارات: {e}")

    # كشف الرجوع بعد انقطاع: لو كان منقطعاً ورجع يرسل، نُشعر العميل
    try:
        if kiln.was_online == 0:
            from app.core.notifications import send_notification
            send_notification(kiln, "✅ عاد الفرن للاتصال بالإنترنت ويرسل القراءات الآن.", title="عودة الاتصال", db=db)
        if kiln.was_online != 1:
            kiln.was_online = 1
            db.commit()
    except Exception as e:
        print(f"تنبيه: تعذّر تحديث حالة الاتصال: {e}")

    return saved_reading if saved_reading else reading


# ═════════════ الإيقاف الإجباري ═════════════

@router.post("/kilns/{kiln_id}/stop", status_code=200)
def request_stop(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """المستخدم يرفع علم الإيقاف. الأردوينو يسأل عنه ويتوقف."""
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    kiln.stop_requested = 1
    db.commit()
    log_event(db, kiln.id, "stop", "🛑 إيقاف إجباري",
              message="قام المستخدم بالضغط على زر الإيقاف الإجباري لإيقاف الفرن.",
              color="#ff5252", icon="🛑")
    return {"status": "ok", "message": "تم إرسال أمر الإيقاف"}


@router.get("/device/stop-status", response_model=StopStatusResponse)
def device_check_stop(kiln: Kiln = Depends(get_kiln_by_device_key), db: Session = Depends(get_db)):
    """
    الأردوينو يسأل: هل فيه أمر إيقاف؟
    نمط "الاستهلاك لمرة واحدة": لو العلم مرفوع، نرجّعه true ثم ننزّله فوراً
    في نفس اللحظة. هكذا يُسلّم الأمر مرة واحدة فقط ولا تحدث حلقة تكرار،
    دون الاعتماد على تأكيد منفصل من الأردوينو قد يفشل لأسباب شبكية.
    """
    was_requested = bool(kiln.stop_requested)
    if was_requested:
        kiln.stop_requested = 0   # ننزّل العلم فور تسليمه
        db.commit()
    return StopStatusResponse(stop_requested=was_requested)


@router.post("/device/stop-confirm", status_code=200)
def device_confirm_stop(kiln: Kiln = Depends(get_kiln_by_device_key), db: Session = Depends(get_db)):
    """الأردوينو يؤكّد أنه نفّذ التوقف؛ هنا فقط ننزّل العلم."""
    kiln.stop_requested = 0
    db.commit()
    return {"status": "ok"}


# ═════════════ إعدادات الإشعارات ═════════════

@router.get("/kilns/{kiln_id}/notify", response_model=NotifySettingsResponse)
def get_notify(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    return NotifySettingsResponse(
        notify_channel=kiln.notify_channel,
        notify_enabled=bool(kiln.notify_enabled),
        notify_interval=kiln.notify_interval,
        stage_notify=bool(kiln.stage_notify),
        pushover_configured=bool(kiln.pushover_token and kiln.pushover_user),
        telegram_configured=bool(kiln.telegram_token and kiln.telegram_chat),
    )


@router.put("/kilns/{kiln_id}/notify", response_model=NotifySettingsResponse)
def set_notify(kiln_id: str, body: NotifySettings, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    data = body.dict(exclude_unset=True)
    # حقول منطقية تُحوّل لأرقام
    if "notify_enabled" in data: kiln.notify_enabled = 1 if data["notify_enabled"] else 0
    if "stage_notify"   in data: kiln.stage_notify   = 1 if data["stage_notify"] else 0
    if "notify_interval" in data: kiln.notify_interval = int(data["notify_interval"])
    if "notify_channel"  in data: kiln.notify_channel  = data["notify_channel"]
    # التوكنات (تُحفظ كما هي)
    for f in ("pushover_token", "pushover_user", "telegram_token", "telegram_chat"):
        if f in data:
            setattr(kiln, f, data[f] or None)
    kiln.last_notified_temp = 0  # إعادة ضبط العدّاد عند تغيير الإعدادات
    db.commit(); db.refresh(kiln)
    return get_notify(kiln_id, current_user, db)


@router.post("/kilns/{kiln_id}/notify/test", status_code=200)
def test_notify(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    kiln = _get_owned_kiln(kiln_id, current_user, db)
    sent = send_notification(kiln, "🔔 هذا إشعار تجريبي من منصة الأفران 🌡️", title="تجريبي", db=db)
    if not sent:
        raise HTTPException(status_code=400, detail="لم تُضبط توكنات القناة المختارة")
    return {"status": "ok", "message": "تم إرسال الإشعار التجريبي"}


# ═════════════ سجل الأحداث ═════════════

@router.get("/kilns/{kiln_id}/events", response_model=List[EventResponse])
def get_kiln_events(kiln_id: str, limit: int = 500, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """أحداث الفرن المهمة (الأحدث أولاً) — لصفحة سجل الأحداث."""
    _get_owned_kiln(kiln_id, current_user, db)
    limit = max(1, min(limit, 2000))
    return (
        db.query(Event).filter(Event.kiln_id == kiln_id)
        .order_by(Event.created_at.desc()).limit(limit).all()
    )


@router.get("/kilns/{kiln_id}/export.csv")
def export_readings_csv(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """تحميل كل القراءات كملف CSV (يفتح في Excel، يدعم العربي)."""
    import csv, io
    from fastapi.responses import StreamingResponse

    kiln = _get_owned_kiln(kiln_id, current_user, db)
    rows = (
        db.query(Reading).filter(Reading.kiln_id == kiln_id)
        .order_by(Reading.recorded_at.asc()).all()
    )

    H_NAMES = {0: "متوقف", 1: "المؤقت", 2: "الحرق التصاعدي", 3: "التثبيت", 4: "النزول التدريجي", 5: "انتهى"}
    columns = [
        ("recorded_at", "الوقت"), ("c1", "حرارة حقيقية"), ("i1", "حرارة افتراضية"),
        ("x", "الدرجة النهائية"), ("h", "الساعات"), ("H", "حالة البرنامج"),
        ("MARAHEL", "المراحل"), ("DOWN", "النزول التدريجي"),
        ("ElectricOff", "حالة الحساس"), ("wiresActive", "الأسلاك"),
    ]

    out = io.StringIO()
    out.write("\ufeff")  # BOM عشان Excel يقرأ العربي
    writer = csv.writer(out)
    writer.writerow([ar for (_, ar) in columns])
    for r in rows:
        row = []
        for (key, _) in columns:
            val = getattr(r, key, "")
            if key == "recorded_at" and val:
                val = _local_time(val, current_user.timezone or 'Asia/Muscat')
            elif key == "H":
                val = H_NAMES.get(val, val)
            elif key == "ElectricOff":
                val = "خلل" if val == 1 else "يعمل"
            elif key in ("MARAHEL", "DOWN"):
                val = "مفعّل" if val == 1 else "غير مفعّل"
            row.append("" if val is None else val)
        writer.writerow(row)

    out.seek(0)
    safe_name = (kiln.name or "kiln").replace(" ", "_")
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_readings.csv"'},
    )


@router.get("/kilns/{kiln_id}/events.csv")
def export_events_csv(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """تحميل سجل الأحداث (تحوّل المراحل، الإشعارات، الإيقاف) كملف CSV يفتح في Excel."""
    import csv, io
    from fastapi.responses import StreamingResponse
    from app.models.kiln import Event

    kiln = _get_owned_kiln(kiln_id, current_user, db)
    rows = (
        db.query(Event).filter(Event.kiln_id == kiln_id)
        .order_by(Event.created_at.asc()).all()
    )

    type_ar = {"stage": "تحوّل مرحلة", "notification": "إشعار حرارة", "stop": "إيقاف إجباري"}

    out = io.StringIO()
    out.write("\ufeff")  # BOM عشان Excel يقرأ العربي
    writer = csv.writer(out)
    writer.writerow(["الوقت", "النوع", "العنوان", "التفاصيل"])
    for ev in rows:
        t = _local_time(ev.created_at, current_user.timezone or 'Asia/Muscat')
        writer.writerow([
            t,
            type_ar.get(ev.type, ev.type),
            ev.title or "",
            ev.message or "",
        ])

    out.seek(0)
    safe_name = (kiln.name or "kiln").replace(" ", "_")
    return StreamingResponse(
        iter([out.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_events.csv"'},
    )


@router.get("/kilns/{kiln_id}/export.xlsx")
def export_unified_xlsx(kiln_id: str, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    تقرير Excel موحّد في ورقة واحدة: كل القراءات والأحداث متسلسلة زمنياً.
    - صفوف القراءات: 21 عموداً بكل المتغيرات، ملوّنة حسب المرحلة.
    - صفوف الأحداث: العمود A فيه الوقت، والباقي مدموج فيه نص الحدث بلون بارز.
    """
    import io
    from datetime import datetime
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from app.models.kiln import Event

    kiln = _get_owned_kiln(kiln_id, current_user, db)
    tz = current_user.timezone or "Asia/Muscat"

    H_NAMES = {0: "متوقف", 1: "المؤقت", 2: "الحرق التصاعدي", 3: "التثبيت", 4: "النزول التدريجي", 5: "انتهى"}
    # ألوان خلفية القراءات حسب المرحلة (فاتحة)
    STAGE_FILL = {
        "متوقف": "FFF0F0F0", "المؤقت": "FFDCEEFF", "الحرق التصاعدي": "FFFFE0D0",
        "التثبيت": "FFFFDADA", "النزول التدريجي": "FFD4F4F9", "انتهى": "FFDCF5DE", "—": "FFFFFFFF",
    }
    # ألوان صفوف الأحداث (غامقة/بارزة) حسب نوع الحدث
    EVENT_FILL = {"stage": "FFFF7043", "notification": "FFFFB74D", "stop": "FFEF5350"}
    EVENT_TEXT = "FFFFFFFF"

    thin = Side(style="thin", color="FFE0E0E0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FF2D2D3A")
    header_font = Font(name="Arial", bold=True, color="FFFFFFFF", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "سجل الفرن"
    ws.sheet_view.rightToLeft = True

    headers = [
        "الوقت والتاريخ", "الحرارة الحقيقية", "الحرارة الافتراضية", "الدرجة النهائية",
        "ساعات الحرق الجارية", "دقائق الحرق الجارية", "وقت المرحلة الواحدة", "مدة التثبيت", "دقائق التثبيت الجارية",
        "ساعات متبقية", "دقائق متبقية", "حالة البرنامج", "المراحل", "النزول التدريجي",
        "حالة الأسلاك", "حالة الحساس",
        "حرارة المرحلة 1", "وقت المرحلة 1", "حرارة المرحلة 2", "وقت المرحلة 2",
        "حرارة المرحلة 3", "وقت المرحلة 3",
    ]
    NCOL = len(headers)  # 22

    # ───── العنوان الكبير ─────
    ws.merge_cells(f"A1:{get_column_letter(NCOL)}1")
    created = _local_time(datetime.utcnow(), tz)[:16]
    tcell = ws["A1"]
    tcell.value = f"📋 سجل الفرن الكامل — القراءات والأحداث المهمة   (أُنشئ في {created})"
    tcell.font = Font(name="Arial", bold=True, size=14, color="FFFF2D75")
    tcell.alignment = center
    ws.row_dimensions[1].height = 30

    # ───── الترويسات ─────
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border
    ws.row_dimensions[2].height = 34

    # ───── جمع القراءات والأحداث ودمجها زمنياً ─────
    readings = db.query(Reading).filter(Reading.kiln_id == kiln_id).order_by(Reading.recorded_at.asc()).all()
    events = db.query(Event).filter(Event.kiln_id == kiln_id).order_by(Event.created_at.asc()).all()

    merged = []
    for rd in readings:
        merged.append((rd.recorded_at, "reading", rd))
    for ev in events:
        merged.append((ev.created_at, "event", ev))
    # ترتيب زمني موحّد (نتعامل مع None بأمان)
    merged.sort(key=lambda x: (x[0] is None, x[0]))

    def onoff(v, yes="مفعّل", no="غير مفعّل"):
        return yes if v == 1 else no

    r = 3
    for ts, kind, obj in merged:
        if kind == "reading":
            rd = obj
            stage = H_NAMES.get(rd.H, "—")
            vals = [
                _local_time(rd.recorded_at, tz),
                rd.c1, rd.i1, rd.x, rd.h, rd.m, rd.t, rd.D, rd.mD, rd.ht, rd.mt,
                stage,
                onoff(rd.MARAHEL), onoff(rd.DOWN),
                "تعمل" if rd.wiresActive not in (None, "", "0", 0) else "متوقفة",
                "عطل" if rd.ElectricOff == 1 else "يعمل",
                rd.x1, rd.t1, rd.x2, rd.t2, rd.x3, rd.t3,
            ]
            fill = STAGE_FILL.get(stage, "FFFFFFFF")
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=v if v is not None else 0)
                c.fill = PatternFill("solid", fgColor=fill)
                c.font = Font(name="Arial", size=9, color="FF333333", bold=(col == 11))
                c.alignment = center; c.border = border
        else:
            ev = obj
            # العمود A: الوقت | الباقي مدموج: نص الحدث
            tcell = ws.cell(row=r, column=1, value=_local_time(ev.created_at, tz)[11:19])
            efill = EVENT_FILL.get(ev.type, "FF888888")
            tcell.fill = PatternFill("solid", fgColor=efill)
            tcell.font = Font(name="Arial", size=9, bold=True, color=EVENT_TEXT)
            tcell.alignment = center; tcell.border = border

            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NCOL)
            msg = f"{ev.icon or ''}  {ev.title or ''}"
            if ev.message:
                msg += f"  —  {ev.message}"
            ecell = ws.cell(row=r, column=2, value=msg)
            ecell.fill = PatternFill("solid", fgColor=efill)
            ecell.font = Font(name="Arial", size=10, bold=True, color=EVENT_TEXT)
            ecell.alignment = Alignment(horizontal="right", vertical="center")
            # حدود لكل خلايا الدمج
            for col in range(2, NCOL + 1):
                ws.cell(row=r, column=col).border = border
        r += 1

    # ───── عرض الأعمدة + تجميد ─────
    widths = [20, 13, 14, 12, 12, 13, 14, 10, 14, 11, 11, 14, 11, 13, 12, 11, 13, 12, 13, 12, 13, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = (kiln.name or "kiln").replace(" ", "_")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_report.xlsx"'},
    )


@router.get("/kilns/{kiln_id}/timeline")
def get_timeline(kiln_id: str, period: str = "day", current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """
    السجل الموحّد: قراءات + أحداث متسلسلة زمنياً، مع فلتر بالفترة.
    period: day | week | month | all
    يُرجّع عناصر مرتّبة زمنياً، كل عنصر إما 'reading' أو 'event'.
    """
    from datetime import datetime, timedelta
    from app.models.kiln import Event

    kiln = _get_owned_kiln(kiln_id, current_user, db)
    tz = current_user.timezone or "Asia/Muscat"

    # نطاق الفترة
    now = datetime.utcnow()
    ranges = {"day": timedelta(days=1), "week": timedelta(days=7), "month": timedelta(days=30)}
    since = None if period == "all" else now - ranges.get(period, ranges["day"])

    rq = db.query(Reading).filter(Reading.kiln_id == kiln_id)
    eq = db.query(Event).filter(Event.kiln_id == kiln_id)
    if since is not None:
        rq = rq.filter(Reading.recorded_at >= since)
        eq = eq.filter(Event.created_at >= since)

    readings = rq.order_by(Reading.recorded_at.asc()).all()
    events = eq.order_by(Event.created_at.asc()).all()

    H_NAMES = {0: "متوقف", 1: "المؤقت", 2: "الحرق التصاعدي", 3: "التثبيت", 4: "النزول التدريجي", 5: "انتهى"}

    items = []
    for rd in readings:
        items.append({
            "kind": "reading",
            "ts": rd.recorded_at.isoformat() if rd.recorded_at else "",
            "time": _local_time(rd.recorded_at, tz),
            "c1": rd.c1, "i1": rd.i1, "x": rd.x, "h": rd.h,
            "t": rd.t, "D": rd.D, "t1": rd.t1, "t2": rd.t2, "t3": rd.t3, "m": rd.m,
            "mD": rd.mD, "ht": rd.ht, "mt": rd.mt, "x1": rd.x1, "x2": rd.x2, "x3": rd.x3,
            "wiresActive": rd.wiresActive, "ElectricOff": rd.ElectricOff,
            "H": rd.H, "stage": H_NAMES.get(rd.H, "—"),
            "MARAHEL": rd.MARAHEL, "DOWN": rd.DOWN,
        })
    for ev in events:
        items.append({
            "kind": "event",
            "ts": ev.created_at.isoformat() if ev.created_at else "",
            "time": _local_time(ev.created_at, tz),
            "type": ev.type, "title": ev.title, "message": ev.message,
            "color": ev.color, "icon": ev.icon,
        })

    # ترتيب زمني موحّد
    items.sort(key=lambda x: x["ts"])
    return {"kiln_name": kiln.name or "", "period": period, "count": len(items), "items": items}
